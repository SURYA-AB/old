from django.shortcuts import render, redirect
from django.http import HttpResponse
from todolist_app.models import TaskList
from todolist_app.models import Gstworker
from todolist_app.forms import TaskForm
from django.contrib import messages
from django.core.paginator import Paginator
from todolist_app.functions.functions import handle_uploaded_file 
from todolist_app.forms import StudentForm
import csv
import json
#import xlsxwriter
from datetime import datetime
from datetime import timedelta
#from xlsxwriter import Workbook
from openpyxl import Workbook
from zipfile import ZipFile
import zipfile




# Create your views here.
def todolist(request):
          if request.method == "POST":
               form = TaskForm(request.POST or None)
               if form.is_valid():
                    form.save()
               messages.success(request,("New Task Added Successfully!"))     
               return  redirect('todolist')
          else:
               all_tasks = TaskList.objects.all()
               paginator = Paginator(all_tasks, 10)
               page = request.GET.get('pg')
               all_tasks = paginator.get_page(page)
               
               return render(request, 'todolist.html', {'all_tasks': all_tasks}) 
def delete_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.delete()
          return  redirect('todolist')
def edit_task(request, task_id):
     if request.method == "POST":
          if request.method == "POST":
               task = TaskList.objects.get(pk=task_id)
               form = TaskForm(request.POST or None, instance=task)
               if form.is_valid():
                    form.save()
          messages.success(request,("Task Edited "))     
          return  redirect('todolist')
     else:
          task_obj = TaskList.objects.get(pk=task_id)
          return render(request, 'edit.html', {'task_obj': task_obj})
def contact(request):
     context = {
          'Contact_text' : " Welcome to  Contact Page.",
          }
     return render(request, 'contact.html', context)
def complete_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.done = True
          task.save()
          return  redirect('todolist')
def pending_task(request, task_id):
          task = TaskList.objects.get(pk=task_id)
          task.done = False
          task.save()
          return  redirect('todolist')
          
def about(request):
     context = {
          'Welcome_text' : " Welcome About Page.",
          }
     return render(request, 'about.html', context)
def index(request):
     context = {
          'index_text' : " Welcome to  Index Page.",
          }
     return render(request, 'index.html', context)   
def index2(request):  # testing to extract multiple zip files
     if request.method == 'POST':  
          #student = StudentForm(request.POST, request.FILES)
          
          client_file = request.FILES['file']
                # unzip the zip file to the same directory 
          with zipfile.ZipFile(client_file, 'r') as zip_ref:
                    first = zip_ref.infolist()[0]
                    with zip_ref.open(first, "r") as fo:
                        a = json.load(fo)
          return HttpResponse( len(a['b2b']))
     else:  
          student = StudentForm()  
          return render(request,"index2.html",{'form':student})


    # Create your views here.

from django.views.generic.edit import FormView
from .forms import FileFieldForm

class FileFieldView(FormView):
    form_class = FileFieldForm
    template_name = 'index.html'  # Replace with your template.
    #success_url = 'todolist.html' # Replace with your URL or reverse().

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        files = request.FILES.getlist('file_field')
        if form.is_valid():
            for f in files:
                ...  # Do something with each file.
            return self.form_valid(form)
        else:
            return self.form_invalid(form)




def index1a(request):  # Main Code
     if request.method == 'POST':
          form = StudentForm(request.POST, request.FILES)  
          client_file = request.FILES['file']
          files = request.FILES.getlist('file')
      # unzip the zip file to the same directory 
          if form.is_valid(): 
               for f in files:
                    handle_uploaded_file(f)
                    with zipfile.ZipFile(f, 'r') as zip_ref:
                         first = zip_ref.infolist()[0]
                         with zip_ref.open(first, "r") as fo:
                              a = json.load(fo)
          
          #a = json.loads(data)
                         if form.is_valid():  
                              #handle_uploaded_file(request.FILES['file'])
               
                              #return HttpResponse("File uploaded successfuly" ) 
                              response = HttpResponse(
                                   content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                              )
                              response['Content-Disposition'] = 'attachment; filename={date}-CA Ram Report Ver1_0_1.xlsx'.format(
                                   date=datetime.now().strftime('%Y-%m-%d'),
                              )
                              
               
                              
                              workbook = Workbook()
               
               
                              # Get active worksheet/tab
                              worksheet = workbook.active
                              worksheet.title = 'B2B'
                              #worksheet['A1'] = 42
                              #worksheet['A2'] = len(a['b2b'])
                              try:
                                   r_count = 0
                                   i = 0 # stands for number of GSTIN in B2B records
                                   j = 2 # Stands for Row 2 indicates data write is gonna start from 2nd row 
                                   k = 0 # Stands for count of invoices in a GSTIN Record
                                   l = 0 # Stands for number of invoice line items in a invoice record
               
                                   while i < (len(a['b2b'])):
                                        while k < len(a['b2b'][i]['inv']):
                                             while l < (len(a['b2b'][i]['inv'][k]['itms'])):
                                                  #worksheet.cell(row = j, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['num'])
                                                  try:
                                                       worksheet.cell(row = j, column = 8).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 9).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 10).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 11).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['camt'])
                                                  except:
                                                       pass        
                                                  try: 
                                                       worksheet.cell(row = j, column = 12).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['samt'])
                                                  except:
                                                       pass 
                                                  try: 
                                                       worksheet.cell(row = j, column = 13).value = (a['b2b'][i]['inv'][k]['itms'][l]['itm_det']['csamt'])
                                                  except:
                                                       pass   
                                                  try:
                                                       worksheet.cell(row = j, column = 1).value = (a['b2b'][i]['ctin'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 2).value = (a['b2b'][i]['inv'][k]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 3).value = (a['b2b'][i]['inv'][k]['inv_typ'])
                                                  except:   
                                                       pass
                                                  try: 
                                                       worksheet.cell(row = j, column = 4).value = (a['b2b'][i]['inv'][k]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 5).value = (a['b2b'][i]['inv'][k]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 6).value = (a['b2b'][i]['inv'][k]['rchrg'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 7).value = (a['b2b'][i]['inv'][k]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 14).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       worksheet.cell(row = j, column = 15).value = (a['fp'])
                                                  except:
                                                       pass
                                                  
                                                  r_count += 1
                                                  l += 1 # Refers to callout the next invoice level line item hope it starts with 0
                                                  j += 1 # Excel offset move to next row    
                                             l = 0 # Resetting to 0 for a new record 
                                             k += 1 # Refers to callout next invoice item for a gst record
               
                                        i += 1 # Moving to next GSTIN
                                        k = 0 # Resetting to 0 for a new record of Invoice
                              except:
                                   pass
                              #worksheet['A3'] = (a['b2b'][1]['inv'][0]['itms'][0]['num'])
                              worksheet.cell(row = 1, column = 1).value = "Customer GSTIN"
                              worksheet.cell(row = 1, column = 2).value = "Total Invoice Value"
                              worksheet.cell(row = 1, column = 3).value = "Type of Invoice"
                              worksheet.cell(row = 1, column = 4).value = "Place of Supply"
                              worksheet.cell(row = 1, column = 5).value = "Date of Invoice"
                              worksheet.cell(row = 1, column = 6).value = "Rcm Applicable"
                              worksheet.cell(row = 1, column = 7).value = "Invoice Number"
                              worksheet.cell(row = 1, column = 8).value = "Rate"
                              worksheet.cell(row = 1, column = 9).value = "Taxable Value"
                              worksheet.cell(row = 1, column = 10).value = "IGST"
                              worksheet.cell(row = 1, column = 11).value = "CGST"
                              worksheet.cell(row = 1, column = 12).value = "SGST"
                              worksheet.cell(row = 1, column = 13).value = "CESS"
                              
                              ws_Info = workbook.create_sheet("Info") # insert at the end (default)
                              ws_B2CLA = workbook.create_sheet("B2CLA")
                              ws_B2CLA.cell(row = 1, column = 1).value  = "Place of Supply"
                              ws_B2CLA.cell(row = 1, column = 2).value  = "Old Invoice Number"
                              ws_B2CLA.cell(row = 1, column = 3).value  = "Old invoice Date"
                              ws_B2CLA.cell(row = 1, column = 4).value  = "Revised Invoice Number"
                              ws_B2CLA.cell(row = 1, column = 5).value  = "Revised Invoice Date"
                              ws_B2CLA.cell(row = 1, column = 6).value  = "Total Invoice Value"
                              ws_B2CLA.cell(row = 1, column = 7).value  = "Applicable per of Diff Tax"
                              ws_B2CLA.cell(row = 1, column = 8).value  = "Invoice Type"
                              ws_B2CLA.cell(row = 1, column = 9).value  = "Taxable Value"
                              ws_B2CLA.cell(row = 1, column = 10).value  = "Rate"
                              ws_B2CLA.cell(row = 1, column = 11).value  = "IGST"
                              ws_B2CLA.cell(row = 1, column = 12).value  = "CGST"
                              ws_B2CLA.cell(row = 1, column = 13).value  = "SGST"
                              ws_B2CLA.cell(row = 1, column = 14).value  = "Cess"
                              
               
               
               
                              
                              
               
                              i = 1
                              for key in a.keys():
                                   if isinstance(a[key], dict)== False:
                                        ws_Info.cell(row = i, column = 1).value = (key) 
                                       # ws_Info.cell(row = i, column = 2).value = len(a[key])  
                                        i += 1
                              ws_B2CL = workbook.create_sheet("B2CL")
                              ws_B2CL.cell(row = 1, column = 1).value  = "Invoice Number"
                              ws_B2CL.cell(row = 1, column = 2).value  = "Date of Invoice"
                              ws_B2CL.cell(row = 1, column = 3).value  = "Total Invoice Value"
                              ws_B2CL.cell(row = 1, column = 4).value  = "Place of Supply"
                              ws_B2CL.cell(row = 1, column = 5).value  = "Rate"
                              ws_B2CL.cell(row = 1, column = 6).value  = "Taxable Value"
                              ws_B2CL.cell(row = 1, column = 7).value  = "IGST"
                              ws_B2CL.cell(row = 1, column = 8).value  = "CGST"
                              ws_B2CL.cell(row = 1, column = 9).value  = "SGST"
                              ws_B2CL.cell(row = 1, column = 10).value  = "CESS"
                              ws_B2CL.cell(row = 1, column = 11).value  = "Applicable ""%" "of TAx rate"
                              ws_B2CL.cell(row = 1, column = 12).value  = "Section 7 Supplies"
                              ws_B2CL.cell(row = 1, column = 13).value  = "GSTIN OF Dealer"
                              ws_B2CL.cell(row = 1, column = 14).value  = "Filing Period"
                              
               
               
                                   
                                   
                              try:
                                   sa = 0
                                   j = 2
                                   while sa < (len(a['b2cl'])):
                                                               
                                        sb = 0    
                                        while sb < (len(a['b2cl'][sa]['inv'])):
                                             sc = 0
                                             while sc < (len(a['b2cl'][sa]['inv'][0]['itms'])):
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 5).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 6).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 7).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 8).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['camt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 9).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 10).value  = (a['b2cl'][sa]['inv'][sb]['itms'][sc]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 1).value = (a['b2cl'][sa]['inv'][sb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 2).value = (a['b2cl'][sa]['inv'][sb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 3).value = (a['b2cl'][sa]['inv'][sb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 12).value = (a['b2cl'][sa]['inv'][sb]['inv_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 11).value = (a['b2cl'][sa]['inv'][sb]['diff_percent'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 4).value = (a['b2cl'][sa]['pos'])  
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 13).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CL.cell(row = j, column = 14).value = (a['fp'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  j += 1
                                                  sc += 1
                                             sb += 1
                                             sc = 0
                                        sb = 0
                                        sc = 0         
                                        sa += 1 
                              except:
                                   pass
                              try:                  
                                   ws_B2BA = workbook.create_sheet("B2BA")
                                   ws_B2BA.cell(row = 1, column = 1).value  = "Customer GSTIN"
                                   ws_B2BA.cell(row = 1, column = 2).value  = "Old Invoice Number" 
                                   ws_B2BA.cell(row = 1, column = 3).value  = "Old Invoice Date" 
                                   ws_B2BA.cell(row = 1, column = 4).value  = "Invoice Number" 
                                   ws_B2BA.cell(row = 1, column = 5).value  = "Invoice Date" 
                                   ws_B2BA.cell(row = 1, column = 6).value  = "Total Invoice Value" 
                                   ws_B2BA.cell(row = 1, column = 7).value  = "Place of Supply" 
                                   ws_B2BA.cell(row = 1, column = 8).value  = "RCM Applicability" 
                                   ws_B2BA.cell(row = 1, column = 9).value  = "Applicable ""%" "of Tax rate" 
                                   ws_B2BA.cell(row = 1, column = 10).value  = "Invoice Type" 
                                   ws_B2BA.cell(row = 1, column = 11).value  = "Taxable Value" 
                                   ws_B2BA.cell(row = 1, column = 12).value  = "Rate" 
                                   ws_B2BA.cell(row = 1, column = 13).value  = "IGST" 
                                   ws_B2BA.cell(row = 1, column = 14).value  = "CGST" 
                                   ws_B2BA.cell(row = 1, column = 15).value  = "SGST" 
                                   ws_B2BA.cell(row = 1, column = 16).value  = "CESS"
                                   ws_B2BA.cell(row = 1, column = 17).value  = "GSTIN OF Client"
                                   ws_B2BA.cell(row = 1, column = 18).value  = "Filing Period"         
                              except:
                                   pass
                              
                              #try:
                              ta = 0
                              j = 2
                              try:
                                   while ta < (len(a['b2ba'])):
                                        tb = 0
                                        while tb < (len(a['b2ba'][ta]['inv'])):
                                             tc = 0
                                             while tc < (len(a['b2ba'][ta]['inv'][tb]['itms'])):
                                                  try: 
                                                       ws_B2BA.cell(row = j, column = 1).value  = ((a['b2ba'][ta]['ctin']))
                                                  except: 
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 2).value  = ((a['b2ba'][ta]['inv'][tb]['oinum']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 3).value  = ((a['b2ba'][ta]['inv'][tb]['oidt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 4).value  = ((a['b2ba'][ta]['inv'][tb]['inum']))
                                                  except:
                                                       pass
                                                  try :
                                                       ws_B2BA.cell(row = j, column = 5).value  = ((a['b2ba'][ta]['inv'][tb]['idt']))
                                                  except:  
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 6).value  = ((a['b2ba'][ta]['inv'][tb]['val']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 7).value  = ((a['b2ba'][ta]['inv'][tb]['pos']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 8).value  = ((a['b2ba'][ta]['inv'][tb]['rchrg']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 9).value  = ((a['b2ba'][ta]['inv'][tb]['diff_percent']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 10).value  = ((a['b2ba'][ta]['inv'][tb]['inv_typ']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 11).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['txval']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 12).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['rt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 13).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['iamt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 14).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['camt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 15).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['samt']))
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 16).value  = ((a['b2ba'][ta]['inv'][tb]['itms'][tc]['itm_det']['csamt']))              
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 17).value = (a['gstin'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2BA.cell(row = j, column = 18).value = (a['fp'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  j += 1
                                                  tc += 1
                                             tb += 1
                                        ta += 1          
                                  
                                  
                              except:
                                   pass
                              
                              ws_Info.cell(row = 4, column = 3).value = "B2CLA Count"
                              #ws_Info.cell(row = 4, column = 4).value = (len(a['b2cla']))    
                              try: 
                                   ua = 0
                                   j = 2
                                   while ua < (len(a['b2cla'])):                         
                                        ub = 0
                                        while ub < (len(a['b2cla'][ua]['inv'])):
                                             uc = 0   
                                             while uc < (len(a['b2cla'][ua]['inv'][ub]['itms'])):
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 1).value  = (a['b2cla'][ua]['pos'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 9).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 10).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 11).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 12).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['camt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 13).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['samt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 14).value  = (a['b2cla'][ua]['inv'][ub]['itms'][uc]['itm_det']['csamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 2).value  = (a['b2cla'][ua]['inv'][ub]['oinum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 3).value  = (a['b2cla'][ua]['inv'][ub]['oidt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 4).value  = (a['b2cla'][ua]['inv'][ub]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 5).value  = (a['b2cla'][ua]['inv'][ub]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 6).value  = (a['b2cla'][ua]['inv'][ub]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 7).value  = (a['b2cla'][ua]['inv'][ub]['diff_percent'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_B2CLA.cell(row = j, column = 8).value  = (a['b2cla'][ua]['inv'][ub]['inv_typ'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  
                                                  
                                                  
                                                  
               
               
               
               
                                                  j += 1
                                                  uc += 1
                                             ub += 1
                                        ua += 1  
                              except:
                                   pass        
                              ws_B2CS = workbook.create_sheet("B2CS")
                              ws_B2CS.cell(row = 1, column = 1).value  = "Supply Type"
                              ws_B2CS.cell(row = 1, column = 2).value  = "Rate"
                              ws_B2CS.cell(row = 1, column = 3).value  = "Suppy is E- Commerce"
                              ws_B2CS.cell(row = 1, column = 4).value  = "Place of Supply"
                              ws_B2CS.cell(row = 1, column = 5).value  = "Differential Tax Rate"   
                              ws_B2CS.cell(row = 1, column = 6).value  = "Taxable Value"
                              ws_B2CS.cell(row = 1, column = 7).value  = "IGST"
                              ws_B2CS.cell(row = 1, column = 8).value  = "CGST"
                              ws_B2CS.cell(row = 1, column = 9).value  = "SGST"
                              ws_B2CS.cell(row = 1, column = 10).value  = "CESS"
                              
                              
                              
               
                              try:
                                   va = 0
                                   j = 2
                                   while va < (len(a['b2cs'])):
                                        try:
                                             ws_B2CS.cell(row = j, column = 1).value  = (a['b2cs'][va]['sply_ty'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 2).value  = (a['b2cs'][va]['rt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 3).value  = (a['b2cs'][va]['typ'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 4).value  = (a['b2cs'][va]['pos'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 5).value  = (a['b2cs'][va]['diff_percent'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 6).value  = (a['b2cs'][va]['txval'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 7).value  = (a['b2cs'][va]['iamt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 8).value  = (a['b2cs'][va]['camt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 9).value  = (a['b2cs'][va]['samt'])
                                        except:
                                             pass
                                        try:
                                             ws_B2CS.cell(row = j, column = 10).value  = (a['b2cs'][va]['csamt'])
                                        except:
                                             pass
                                        va += 1
                                        j += 1
                              except:
                                   pass
                              
                              
                                        
                                        
                              ws_B2CSA = workbook.create_sheet("B2CSA")
                              ws_B2CSA.cell(row = 1, column = 1).value  = "Original Month"
                              ws_B2CSA.cell(row = 1, column = 2).value  = "Supply Type"
                              ws_B2CSA.cell(row = 1, column = 3).value  = "Ecomerce ?"
                              ws_B2CSA.cell(row = 1, column = 4).value  = "Place of Supply"
                              ws_B2CSA.cell(row = 1, column = 5).value  = "Differential Tax Rate"   
                              ws_B2CSA.cell(row = 1, column = 6).value  = "Taxable Value"
                              ws_B2CSA.cell(row = 1, column = 7).value  = "IGST"
                              ws_B2CSA.cell(row = 1, column = 8).value  = "CGST"
                              ws_B2CSA.cell(row = 1, column = 9).value  = "SGST"
                              ws_B2CSA.cell(row = 1, column = 10).value  = "CESS"
               
               
               
                              try:
                                   wa = 0
                                   j = 2
                                   while wa < (len(a['b2csa'])):
                                        wb = 0
                                        while wb < (len(a['b2csa'][wa]['itms'])):
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 1).value  = (a['b2csa'][wa]['omon'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 2).value  = (a['b2csa'][wa]['sply_ty'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 3).value  = (a['b2csa'][wa]['typ'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 4).value  = (a['b2csa'][wa]['pos'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 5).value  = (a['b2csa'][wa]['diff_percent'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 6).value  = (a['b2csa'][wa]['itms'][wb]['txval'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 7).value  = (a['b2csa'][wa]['itms'][wb]['rt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 8).value  = (a['b2csa'][wa]['itms'][wb]['iamt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 9).value  = (a['b2csa'][wa]['itms'][wb]['camt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 10).value  = (a['b2csa'][wa]['itms'][wb]['samt'])
                                             except:
                                                  pass
                                             try:
                                                  ws_B2CSA.cell(row = j, column = 11).value  = (a['b2csa'][wa]['itms'][wb]['csamt'])
                                             except:
                                                  pass
                                             r_count += 1
                                             j += 1
                                             wb += 1
                                        wa += 1
                              except:
                                   pass
                              
                                       
                              ws_EXP = workbook.create_sheet("EXP")
                              ws_EXP.cell(row = 1, column = 1).value  = "Export Type"
                              ws_EXP.cell(row = 1, column = 2).value  = "Invoice Number"
                              ws_EXP.cell(row = 1, column = 3).value  = "Invoice Date"
                              ws_EXP.cell(row = 1, column = 4).value  = "Invoice Value"
                              ws_EXP.cell(row = 1, column = 5).value  = "Port Code"
                              ws_EXP.cell(row = 1, column = 6).value  = "Shipping Bill Number"
                              ws_EXP.cell(row = 1, column = 7).value  = "Shipping Bill Date"
                              ws_EXP.cell(row = 1, column = 8).value  = "Taxable Value"
                              ws_EXP.cell(row = 1, column = 9).value  = "Rate"
                              ws_EXP.cell(row = 1, column = 10).value  = "IGST"
                              ws_EXP.cell(row = 1, column = 11).value  = "CESS"
               
               
               
                                        
                              j = 2
                              xa = 0
                              try:
                                   while xa < (len(a['exp'])):
                                        xb = 0
                                        while xb < (len(a['exp'][xa]['inv'])):
                                             xc = 0
                                             while xc < (len(a['exp'][xa]['inv'][xb]['itms'])):
                                                  try:
                                                       ws_EXP.cell(row = j, column = 1).value  = (a['exp'][xa]['exp_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 2).value  = (a['exp'][xa]['inv'][xb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 3).value  = (a['exp'][xa]['inv'][xb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 4).value  = (a['exp'][xa]['inv'][xb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 5).value  = (a['exp'][xa]['inv'][xb]['sbpcode'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 6).value  = (a['exp'][xa]['inv'][xb]['sbnum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 7).value  = (a['exp'][xa]['inv'][xb]['sbdt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 8).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 9).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 10).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXP.cell(row = j, column = 11).value  = (a['exp'][xa]['inv'][xb]['itms'][xc]['csamt'])
                                                  except:
                                                       pass
                                                  r_count += 1
                                                  j += 1
                                                  xc += 1
                                             xb += 1
                                        xa += 1
                              except:
                                   pass   
                              
                              ws_EXPA = workbook.create_sheet("EXPA")
                              ws_EXPA.cell(row = 1, column = 1).value  = "Export Type"
                              ws_EXPA.cell(row = 1, column = 2).value  = "Invoice Number"
                              ws_EXPA.cell(row = 1, column = 3).value  = "Invoice Date"
                              ws_EXPA.cell(row = 1, column = 4).value  = "Invoice Value"
                              ws_EXPA.cell(row = 1, column = 5).value  = "Port Code"
                              ws_EXPA.cell(row = 1, column = 6).value  = "Shipping Bill Number"
                              ws_EXPA.cell(row = 1, column = 7).value  = "Shipping Bill Date"
                              ws_EXPA.cell(row = 1, column = 8).value  = "Taxable Value"
                              ws_EXPA.cell(row = 1, column = 9).value  = "Rate"
                              ws_EXPA.cell(row = 1, column = 10).value  = "IGST"
                              ws_EXPA.cell(row = 1, column = 11).value  = "CESS"  
                              ws_EXPA.cell(row = 1, column = 12).value  = "Old Invoice Number"  
                              ws_EXPA.cell(row = 1, column = 13).value  = "Old Invoice Date"  
               
                              j = 2
                              ya = 0
                              try:
                                   while ya < (len(a['expa'])):
                                        yb = 0
                                        while yb < (len(a['expa'][ya]['inv'])):
                                             yc = 0
                                             while yc < (len(a['expa'][ya]['inv'][yb]['itms'])):
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 1).value  = (a['expa'][ya]['exp_typ'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 2).value  = (a['expa'][ya]['inv'][yb]['inum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 3).value  = (a['expa'][ya]['inv'][yb]['idt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 4).value  = (a['expa'][ya]['inv'][yb]['val'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 5).value  = (a['expa'][ya]['inv'][yb]['sbpcode'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 6).value  = (a['expa'][ya]['inv'][yb]['sbnum'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 7).value  = (a['expa'][ya]['inv'][yb]['sbdt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 8).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['txval'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 9).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['rt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 10).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['iamt'])
                                                  except:
                                                       pass
                                                  try:
                                                       ws_EXPA.cell(row = j, column = 11).value  = (a['expa'][ya]['inv'][yb]['itms'][yc]['csamt'])
                                                  except:
                                                       pass
                                                  ws_EXPA.cell(row = j, column = 12).value  = (a['expa'][ya]['inv'][yb]['oinum'])
                                                  ws_EXPA.cell(row = j, column = 13).value  = (a['expa'][ya]['inv'][yb]['oidt'])
                                                  r_count += 1
                                                  j += 1
                                                  yc += 1
                                             yb += 1
                                        ya += 1
                              except:
                                   pass                         
                              
                                   
                              ws_HSN = workbook.create_sheet("HSN")
                              ws_HSN.cell(row = 1, column = 1).value  = "S No"
                              ws_HSN.cell(row = 1, column = 2).value  = "HSN Code"
                              ws_HSN.cell(row = 1, column = 3).value  = "Description"   
                              
                              za = 0
                              j = 2
                              try:
                                   while za < (len(a['hsn']['data'])): 
                                        try: 
                                             ws_HSN.cell(row = j, column = 1).value  = (a['hsn']['data'][za]['num'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = j, column = 2).value  = (a['hsn']['data'][za]['hsn_sc'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = j, column = 3).value  = (a['hsn']['data'][za]['desc'])
                                        except:
                                             pass
                                        try:     
                                             ws_HSN.cell(row = j, column = 4).value  = (a['hsn']['data'][za]['uqc'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = j, column = 5).value  = (a['hsn']['data'][za]['qty'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = j, column = 6).value  = (a['hsn']['data'][za]['val'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = j, column = 7).value  = (a['hsn']['data'][za]['txval'])
                                        except:
                                             pass
                                        try:                                   
                                             ws_HSN.cell(row = j, column = 8).value  = (a['hsn']['data'][za]['iamt'])
                                        except:
                                             pass
                                        try:                                   
                                             ws_HSN.cell(row = j, column = 9).value  = (a['hsn']['data'][za]['samt'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = j, column = 10).value  = (a['hsn']['data'][za]['camt'])
                                        except:
                                             pass
                                        try:          
                                             ws_HSN.cell(row = j, column = 11).value  = (a['hsn']['data'][za]['csamt'])
                                        except:
                                             pass
                                        
                                        j += 1
                                        r_count +=1
                                        za += 1  
                                         
                              except:
                                   pass
                                        
                                        
                                        
                                        
                                        
                                        
                                        
                                        
                              
                              
                              
                              
                              
                                        
                                        
                                        
                                        
                                        
                              
                              myobject = Gstworker(GSTIN=(a['gstin']), r_counts=(r_count))
                              myobject.save()              
                                   
               
                              
               
                                   
                              workbook.save(response)
                              
               return response
               

    # xlsx_data contains the Excel file
     else:  
          student = StudentForm()  
          return render(request,"index.html",{'form':student})              

     return self.form_valid(form)
